# install OpenPyXL and modules plus numpy
import os
import openpyxl
from openpyxl import Workbook, load_workbook
import numpy
print("Loading Workbook...")

#------------------------------------------
FILENAME = '1983Devices.xlsx'       # <----- Enter Excel File Name Within load_workbook('example.xlsx')
#-----------------------------------------

book = load_workbook(FILENAME)  
sheet = book.active
sheetLength = sheet.max_row
currentApp = "none"
##Data Processing
#normalize application data : For each application, replace all values under it with the application until a new one appears
for x in range(2, sheetLength):
    cell = "A" + str(x)
    cellNext = "A" + str(x+1)
    cellValue = sheet[cell].value
    nextValue = sheet[cellNext].value

    if currentApp == "none":
        currentApp = cellValue

    if (cellValue == None) or (cellValue == " "):  #If Empty, replace with current app
        sheet[cell].value = currentApp
        print( "Cell Empty. Replacing..")
    elif cellValue != None: # if Not Empty, replace current app with cell value
        currentApp = cellValue
        print( "Cell Value:" + cellValue)
    print("CELL: " + cell + "      CURRENT APP:" +currentApp)
book.save(FILENAME)

#scan the list of buildings. for each , check to see if the building is already in the list. 
#if not,add to list and start counting the iPads and applications. If so, add to the ipad count.

def isInList(cell, list):
    result = sheet[cell].value in list
    if result == True:
        result = list.index(sheet[cell].value)
    return result
book = load_workbook(FILENAME)
sheet = book.active
sheetLength = sheet.max_row
buildings = ["Buildings"]
numDevices= ["# Of Devices"]
deviceIDs = []
apps = [[] * 1 for i in range(10000)] #Accessed using buildings[column][row]
for x in range(2, sheetLength):
        buildCell = "O" + str(x)
        deviceCell = "M" + str(x)
        appCell = "A" + str(x)
        #cellValue = sheet[cell].value
        id = isInList(buildCell, buildings)
        device = isInList(deviceCell, deviceIDs)
        app = isInList(appCell, apps[id])

        if id == False: 
            buildings.append(sheet[buildCell].value)
            numDevices.append(1)
            id = len(buildings) - 1
            print("New Building Registered")
            print("ID =" + str(id))
        if device == False:
            deviceIDs.append(sheet[deviceCell].value)
            numDevices[id] = numDevices[id] + 1
            print("New Device Registered")
            print("Number of Devices in Building ID " + str(id) + ": " + str(numDevices[id]))
        if app == False:
            apps[id].append(sheet[appCell].value)
            print("New App Registered")
        print("Cycle " + str(x) + " of " + str(sheetLength))

print(buildings)
print("Amount of Buildings: " + str(len(buildings)))
print(deviceIDs)
print("Amount of Devices: " + str(len(deviceIDs)))
print(apps)
print("Amount of Apps: " + str(len(apps)))

for y in range(1, len(numDevices)): 
    print("Building:" + str(buildings[y]) + " Devices: "  + str(numDevices[y]))
    print(" Apps: ")
    for z in range(1, len(apps[y])):
        print(str(apps[y][z]))
    print("--------------------------")
book.create_sheet("Buildings|Devices|Apps")
sheet = book["Buildings|Devices|Apps"]
for m in range(len(buildings)):
    sheet.append([buildings[m], numDevices[m]])
    for n in range(len(apps[m])):
            column = 67 + n
            print("Column #" + str(column))
            if column > 90:    #if columns go to Z, loop back to AA
                column = chr(65 + ((column - 90) // 26)) + chr(65 + (n % 26))
            else: column = chr(column)
            cell = str(column) + str(m + 1)   # Column + Row 
            #print(type(apps[m][n]))
            print(cell)
            value = apps[m][n]
            sheet[cell].value = str(value)
        

book.save(FILENAME)
print("Saved.")